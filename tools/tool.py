import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


def format_percent(x):
    return "{:.2%}".format(x)


def add_row_to_excel(stat_name, data, sheet_name="备案率", file_name="汇总数据.xlsx"):
    # 检查文件是否存在，不存在则创建新文件
    try:
        wb = load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()

    # 检查指定名称的工作表是否存在，不存在则创建新工作表
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(["统计名称", "实际备案数", "应备案数", "百分率"])
        for i in range(1, 5):
            col = get_column_letter(i)
            ws.column_dimensions[col].width = 30
            ws.cell(row=1, column=i).font = ws.cell(row=2, column=i).font.copy(bold=True)
            ws.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center')
    else:
        ws = wb[sheet_name]

    # 添加新行数据
    new_row = [stat_name] + data
    ws.append(new_row)

    # 设置单元格样式
    row_num = ws.max_row
    for col_num in range(1, 5):
        cell = ws.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col_num == 4:
            cell.number_format = '0.00%'

    # 保存文件并删除空工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(file_name)


def sort_list(data):
    def key_func(x):
        if x.find("丽水") != -1:
            return (0, x)
        elif x.endswith("区") or x.endswith("县") or x.endswith("市"):
            return (1, x[-1], x)
        else:
            return (2, x)

    sorted_data = sorted(data, key=key_func)
    result = []
    i = 0
    while i < len(sorted_data):
        j = i + 1
        while j < len(sorted_data) and key_func(sorted_data[j])[:3] == key_func(sorted_data[i])[:3]:
            j += 1
        if key_func(sorted_data[i]) == (1, sorted_data[i][-1], sorted_data[i]):
            result += sorted(sorted_data[i:j], key=lambda x: x[-1])
        else:
            result += sorted_data[i:j]
        i = j
    return result


def extract_info(cell_value):
    """用于拆分提取的函数"""
    return cell_value.split("/")[4]


def add_new_column(filename, sheet_name, heading):
    """给指定的表格添加新列"""
    wb = load_workbook(filename)
    sheet = wb[sheet_name]

    num_cols = sheet.max_column  # 获取当前的列数
    new_col_num = num_cols + 1  # 新列的列号为当前列数+1
    new_col_header = heading  # 通过参数传递表头

    sheet.insert_cols(new_col_num)  # 在末尾添加新列
    sheet.cell(row=1, column=new_col_num).value = new_col_header  # 设置表头

    col_num = get_col_by_heading(sheet, "建设部门")  # 获取建设部门列的列号

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=col_num).value  # 获取建设部门列的单元格值
        sheet.cell(row=row, column=new_col_num).value = extract_info(cell_value)  # 在新列中写入提取的值

    wb.save(filename)  # 保存表格


def initialize_table(filename, sheet_name):
    # 预处理表格，添加标识列
    add_new_column(filename, sheet_name, heading="统计标识")
    # 打开工作簿
    workbook = openpyxl.load_workbook(filename)

    # 选择指定的工作表
    worksheet = workbook[sheet_name]

    # 循环遍历所有单元格，并将为空的单元格填充为字符UNKNOWN
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                cell.value = 'UNKNOWN'
    # 保存工作簿
    workbook.save(filename)

    return None


def get_col_by_heading(sheet, heading):
    """根据表头名称，获取对应的列号。"""
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for i, col_name in enumerate(row):
            if col_name == heading:
                return i + 1
    raise ValueError(f'Unable to locate column with heading "{heading}"')


def get_col_data_unique(sheet, heading):
    """根据表头名称，获取对应的列数据(去重)。"""
    col_num = get_col_by_heading(sheet, heading)
    data_set = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_set.add(row[col_num - 1])
    data_list = list(data_set)
    return sort_list(data_list)


def calculate_backup_rate(sheet, construction_unit):
    # 获取列号
    col_construction_unit = get_col_by_heading(sheet, '统计标识')
    col_app_status = get_col_by_heading(sheet, '应用状态')
    col_app_type = get_col_by_heading(sheet, '应用类型')
    col_db_level = get_col_by_heading(sheet, '等保级别')
    col_db_register = get_col_by_heading(sheet, '是否等保备案')
    col_db_level_review = get_col_by_heading(sheet, '是否等保定级')

    # 应备案 = 应用状态为非“停用”、“申报中”、“谋划中” & 应用类型非“硬件类系统” & 等保级别非“一级”

    # 已备案 = 应用状态为非“停用”、“申报中”、“非谋划中” & 应用类型非“硬件类系统” & 是否等保定级“是” & 等保级别“二级”、“三级” & 是否等保备案“是”

    # 三级等保良好通过率 = 三级系统测评分数 >= 80分 / 三级系统总数

    # 计算应备案应用数量
    need_backup_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # if row[col_construction_unit - 1] == construction_unit \
        if construction_unit in row[col_construction_unit - 1] \
                and row[col_app_status - 1] not in ["停用", "申报中", "谋划中"] \
                and row[col_app_type - 1] != "硬件类系统" \
                and row[col_db_level - 1] != "一级":
            need_backup_count += 1

    # 计算已备案应用数量
    backup_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # row[col_construction_unit - 1] == construction_unit
        if construction_unit in row[col_construction_unit - 1] \
                and row[col_app_status - 1] not in ["停用", "申报中", "谋划中"] \
                and row[col_app_type - 1] != "硬件类系统" \
                and row[col_db_register - 1] == "是" \
                and row[col_db_level - 1] in ["二级", "三级"] \
                and row[col_db_level_review - 1] == "是":
            backup_count += 1

    # 计算备案率
    if need_backup_count > 0:
        backup_rate = backup_count / need_backup_count
    else:
        backup_rate = 0

    return [backup_count, need_backup_count, format_percent(backup_rate)]


def calculate_level_three_fine_rate(sheet, construction_unit):
    # 获取列号
    col_construction_unit = get_col_by_heading(sheet, '统计标识')
    col_db_level = get_col_by_heading(sheet, '等保级别')
    col_system_score = get_col_by_heading(sheet, '等保测评得分')
    col_app_status = get_col_by_heading(sheet, '应用状态')
    col_app_type = get_col_by_heading(sheet, '应用类型')

    # 计算三级系统数量和等保良好的数量
    level_three_count = 0
    level_three_good_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if construction_unit in row[col_construction_unit - 1] \
                and row[col_app_status - 1] not in ["停用", "申报中", "谋划中"] \
                and row[col_app_type - 1] != "硬件类系统" \
                and row[col_db_level - 1] == '三级':
            level_three_count += 1
            if row[col_system_score - 1] != "UNKNOWN" \
                    and float(row[col_system_score - 1]) >= 80:
                level_three_good_count += 1

    # 计算通过率
    if level_three_count > 0:
        pass_rate = level_three_good_count / level_three_count
    else:
        pass_rate = 0

    return [level_three_good_count, level_three_count, format_percent(pass_rate)]
